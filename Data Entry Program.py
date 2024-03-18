import openpyxl
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from tkinter import ttk, messagebox
import datetime

selected_sheet = None  # Initialize selected_sheet as None initially

def search_info(selected_sheet, search_term):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb[selected_sheet]  # Select the specified sheet
        
        matching_rows = []
        
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            for value in row:
                if value is None:
                    continue
                if isinstance(value, datetime.datetime):
                    value = value.strftime('%m/%d/%Y')
                if search_term.lower() in str(value).lower():
                    matching_rows.append((row_number, row))
                    break
        
        if matching_rows:
            result_text.delete('1.0', tk.END)
            
            column_widths = [8, 30, 15, 12, 30]  # Adjust column widths as needed
            
            result_text.insert(tk.END, f"Row{' '*(column_widths[0]-3)}")
            result_text.insert(tk.END, f"Name{' '*(column_widths[1]-4)}")
            result_text.insert(tk.END, f"Contact No.{' '*(column_widths[2]-11)}")
            result_text.insert(tk.END, f"Birthdate{' '*(column_widths[3]-9)}")
            result_text.insert(tk.END, f"Address\n")
            
            for row_number, row in matching_rows:
                name, contact, birthdate, address = row
                contact = str(contact) if contact is not None else ''
                birthdate = str(birthdate).split()[0] if birthdate is not None else ''
                result_text.insert(tk.END, f"{row_number}{' '*(column_widths[0]-len(str(row_number)))}")
                result_text.insert(tk.END, f"{name.ljust(column_widths[1])}")
                result_text.insert(tk.END, f"{contact.ljust(column_widths[2])}")
                result_text.insert(tk.END, f"{birthdate.ljust(column_widths[3])}")
                result_text.insert(tk.END, f"{address}\n")
        else:
            result_text.delete('1.0', tk.END)
            result_text.insert(tk.END, "No matching rows found.")
    except FileNotFoundError:
        result_text.delete('1.0', tk.END)
        result_text.insert(tk.END, "The specified Excel file was not found.")
    except Exception as e:
        result_text.delete('1.0', tk.END)
        result_text.insert(tk.END, f"An error occurred: {e}")

def select_sheet():
    global selected_sheet  # Declare selected_sheet as global
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet_names = wb.sheetnames
        selected_sheet = sheet_dialog(excel_file, sheet_names)
        return selected_sheet  # Return the selected sheet name
    except FileNotFoundError:
        messagebox.showerror("Error", "The specified Excel file was not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def sheet_dialog(file_path, sheet_names):
    dialog = tk.Toplevel(window)
    dialog.title("Select Sheet")
    
    label = tk.Label(dialog, text="Select a sheet:")
    label.pack()
    
    selected_sheet = tk.StringVar()
    selected_sheet.set(sheet_names[0])
    sheet_menu = ttk.Combobox(dialog, textvariable=selected_sheet, values=sheet_names, state="readonly")
    sheet_menu.pack()
    
    confirm_button = tk.Button(dialog, text="Select", command=lambda: confirm_sheet(file_path, selected_sheet.get(), dialog))
    confirm_button.pack()
    
    dialog.transient(window)
    dialog.grab_set()
    window.wait_window(dialog)
    
    return selected_sheet.get()

def confirm_sheet(file_path, selected_sheet, dialog):
    dialog.destroy()
    search_info(selected_sheet, search_entry.get())


        
def save_data():
    # Create a new window for saving data
    save_window = tk.Toplevel(window)
    save_window.title("Save Data")

    # Name entry
    name_label = tk.Label(save_window, text="Name:")
    name_label.grid(row=0, column=0, padx=5, pady=5)
    name_var = tk.StringVar()
    name_entry = ttk.Entry(save_window, textvariable=name_var, width=50)
    name_entry.grid(row=0, column=1, padx=5, pady=5)

    # Contact No. entry
    contact_label = tk.Label(save_window, text="Contact No.:")
    contact_label.grid(row=1, column=0, padx=5, pady=5)
    contact_var = tk.StringVar()
    contact_entry = ttk.Entry(save_window, textvariable=contact_var, width=50)
    contact_entry.grid(row=1, column=1, padx=5, pady=5)

    # Birthdate entry
    birthdate_label = tk.Label(save_window, text="Birthdate:")
    birthdate_label.grid(row=2, column=0, padx=5, pady=5)
    birthdate_var = tk.StringVar()
    birthdate_entry = ttk.Entry(save_window, textvariable=birthdate_var, width=50)
    birthdate_entry.grid(row=2, column=1, padx=5, pady=5)

    # Address entry
    address_label = tk.Label(save_window, text="Address:")
    address_label.grid(row=3, column=0, padx=5, pady=5)
    address_var = tk.StringVar()
    address_entry = ttk.Entry(save_window, textvariable=address_var, width=50)
    address_entry.grid(row=3, column=1, padx=5, pady=5)

    # Save button in the new window
    save_button = tk.Button(save_window, text="Save Data", command=lambda: save_data_on_window(name_var.get(), contact_var.get(), birthdate_var.get(), address_var.get(), save_window))
    save_button.grid(row=4, column=0, columnspan=2, padx=5, pady=5)

def save_data_on_window(name, contact, birthdate, address, save_window):
    try:
        wb = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    ws = wb.active

    # Append the data to the Excel file
    ws.append([name, contact, birthdate, address])
    
    # Save the Excel file
    wb.save(excel_file)
    
    # Close the window
    save_window.destroy()
    
    # Show success message
    messagebox.showinfo("Success", "Data saved successfully.")
    
    # Refresh the list of names
    refresh_names()

def edit_data():
    selected_name = search_entry.get()
    if not selected_name:
        messagebox.showinfo("No Name Selected", "Please select a name to edit.")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        found = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == selected_name:
                # Open a new window for editing
                edit_window = tk.Toplevel(window)
                edit_window.title("Edit Data")
                
                # Populate the new window with data
                name_label = tk.Label(edit_window, text="Name:")
                name_label.grid(row=0, column=0, padx=5, pady=5)
                name_var_edit = tk.StringVar(value=row[0])
                name_entry_edit = ttk.Entry(edit_window, textvariable=name_var_edit, width=50)
                name_entry_edit.grid(row=0, column=1, padx=5, pady=5)

                contact_label = tk.Label(edit_window, text="Contact No.:")
                contact_label.grid(row=1, column=0, padx=5, pady=5)
                contact_var_edit = tk.StringVar(value=row[1])
                contact_entry_edit = ttk.Entry(edit_window, textvariable=contact_var_edit, width=50)
                contact_entry_edit.grid(row=1, column=1, padx=5, pady=5)

                birthdate_label = tk.Label(edit_window, text="Birthdate:")
                birthdate_label.grid(row=2, column=0, padx=5, pady=5)
                birthdate_var_edit = tk.StringVar(value=row[2])
                birthdate_entry_edit = ttk.Entry(edit_window, textvariable=birthdate_var_edit, width=50)
                birthdate_entry_edit.grid(row=2, column=1, padx=5, pady=5)

                address_label = tk.Label(edit_window, text="Address:")
                address_label.grid(row=3, column=0, padx=5, pady=5)
                address_var_edit = tk.StringVar(value=row[3])
                address_entry_edit = ttk.Entry(edit_window, textvariable=address_var_edit, width=50)
                address_entry_edit.grid(row=3, column=1, padx=5, pady=5)
                
                # Save button in the new window
                save_button_edit = tk.Button(edit_window, text="Save Changes", command=lambda: save_changes(selected_name, name_var_edit.get(), contact_var_edit.get(), birthdate_var_edit.get(), address_var_edit.get(), edit_window))
                save_button_edit.grid(row=4, column=0, columnspan=2, padx=5, pady=5)
                
                found = True
                break
        
        if not found:
            messagebox.showinfo("Not Found", f"Data for {selected_name} not found.")
    except FileNotFoundError:
        messagebox.showerror("Error", "The specified Excel file was not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def save_changes(selected_name, name, contact, birthdate, address, edit_window):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        row_index = 1
        for row in ws.iter_rows(values_only=True):
            if row[0] == selected_name:
                ws.cell(row=row_index, column=1).value = name
                ws.cell(row=row_index, column=2).value = contact
                ws.cell(row=row_index, column=3).value = birthdate
                ws.cell(row=row_index, column=4).value = address
                break
            row_index += 1
        
        wb.save(excel_file)
        edit_window.destroy()
        messagebox.showinfo("Success", "Changes saved successfully.")
    except FileNotFoundError:
        messagebox.showerror("Error", "The specified Excel file was not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def delete_data(name_to_delete):
    try:
        if not name_to_delete:
            messagebox.showinfo("No Name Entered", "Please enter a name before attempting to delete data.")
            return  # Exit the function if no name is entered

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        found = False
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] == name_to_delete:
                ws.delete_rows(row_idx)
                found = True
                break

        if found:
            wb.save(excel_file)
            messagebox.showinfo("Success", f"Data for {name_to_delete} deleted successfully.")
            refresh_names()
        else:
            messagebox.showinfo("Not Found", f"Data for {name_to_delete} not found.")
    except FileNotFoundError:
        messagebox.showerror("Error", "The specified Excel file was not found.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def clear_search():
    search_var.set("")  # Clear the search entry field
    result_text.delete('1.0', tk.END)  # Clear the result display
    refresh_names()  # Refresh autocomplete list

def on_search_change(event):
    search_term = search_var.get().lower()
    matching_names = [n for n in all_names if search_term.lower() in n.lower()]
    search_entry['values'] = matching_names


def select_excel_file():
    global excel_file
    excel_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    refresh_names()  # Refresh the list of names after selecting the Excel file

# Create the main window
window = tk.Tk()
window.title("Data Entry Program")

# Excel file path

excel_file = ''

# Select Excel file button
select_excel_button = tk.Button(window, text="Select Excel File", command=select_excel_file)
select_excel_button.grid(row=0, column=0, padx=5, pady=5, sticky="w")

# Select Sheet button
select_sheet_button = tk.Button(window, text="Select Sheet", command=select_sheet)
select_sheet_button.grid(row=0, column=0, padx=5, pady=5)


def refresh_names():
    global all_names
    all_names = []  # Clear the list before populating
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        all_names = [row[0] for row in ws.iter_rows(values_only=True)]
        search_entry['values'] = all_names  # Update autocomplete list
    except FileNotFoundError:
        pass  # Handle file not found error gracefully


# Create a frame to contain the buttons
button_frame = tk.Frame(window)
button_frame.grid(row=7, column=0, columnspan=4, padx=5, pady=5)

# Save button
save_button = tk.Button(button_frame, text="Save Data", command=save_data)
save_button.grid(row=0, column=0, padx=5)

# Edit button
edit_button = tk.Button(button_frame, text="Edit Data", command=edit_data)
edit_button.grid(row=0, column=1, padx=5)

# Delete button
delete_button = tk.Button(button_frame, text="Delete Data", command=lambda: delete_data(search_entry.get()))
delete_button.grid(row=0, column=2, padx=5)

# Clear search button
clear_button = tk.Button(button_frame, text="Clear Search", command=clear_search)
clear_button.grid(row=0, column=3, padx=5)

# Search Term entry with autocomplete
search_var = tk.StringVar()
search_entry = ttk.Combobox(window, textvariable=search_var, width=50)
search_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")  # Align to the left
search_entry.bind("<KeyRelease>", on_search_change)  # Bind on_search_change to the KeyRelease event of search_entry

# Search button
search_button = tk.Button(window, text="Search", command=lambda: search_info(selected_sheet, search_entry.get()))
search_button.grid(row=0, column=2, padx=(5, 0), pady=5, sticky="w")  # Align to the left with some padding



# Result display
result_text = tk.Text(window)
result_text.grid(row=6, column=0, columnspan=4, sticky="nsew")
window.rowconfigure(6, weight=1)  # Allow row 6 to expand vertically
window.columnconfigure((0, 1, 2, 3), weight=1)  # Allow columns 0 to 3 to expand horizontally


# Run the GUI
window.mainloop()








